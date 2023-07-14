import json
import os
import random

from openpyxl import Workbook
from openpyxl.styles import Font

from datetime import datetime


class SaveCollAndProduct:
    def __init__(self, collect_dict, good_dict):

        self.colums_checker = {}

        self.collect_dict = collect_dict

        self.good_dict = good_dict['result']

        self.colums_harakt = good_dict['name_colums']

        self.colums = ['ID продукта', 'Код', 'Имя продукта', 'Цена', 'Старая цена', 'Комплект главные',
                       'Комплект дополнительные', 'Фото', 'Категория', 'Производитель', 'Тип товара',
                       'Ссылка на сторонний сайт', 'Алгоритм', 'Ед.Измерения', 'ID', 'PARENT_ID', 'Видимость',
                       'Видимость варианта', 'Статус товара', 'Количество', 'Описание', 'Видео', 'Документы',
                       'Гарантия', 'Артикул', 'Страна', 'Производитель', 'Коллекция', 'Цвет', 'Размер', 'Назначение',
                       'Состав коллекции', 'Рисунок', 'Поверхность', 'Ректификат', 'Форма', 'Ширина, см', 'Длина, см',
                       'Толщина, мм', 'Количество в коробке, м2', 'Количество в коробке, шт.', 'Вес коробки, кг']

    @staticmethod
    def save_to_json(filename, good_data):
        filename = f'{filename}.json'

        try:
            with open(filename, 'w', encoding='utf-8') as file:
                json.dump(good_data, file, indent=4, ensure_ascii=False)
        except:
            return False

        return filename

    def write_collections(self, ws, count_def, post):

        ws.cell(row=count_def, column=1).value = ''
        ws.cell(row=count_def, column=2).value = ''
        try:
            full_name = post['full_name']
        except:
            full_name = ''

        try:
            name = post['name']
        except:
            name = ''

        ws.cell(row=count_def, column=3).value = name
        # ws.cell(row=count_def, column=3).value = full_name

        try:
            price = int(post['price'])
        except:
            try:
                price = post['price']
            except:
                price = ''

        ws.cell(row=count_def, column=4).value = price
        ws.cell(row=count_def, column=5).value = ''
        ws.cell(row=count_def, column=6).value = ''
        ws.cell(row=count_def, column=7).value = ''

        try:
            image = ' '.join(x for x in post['image'])
        except:
            image = ''

        ws.cell(row=count_def, column=8).value = image

        try:
            category = 'Плитка, керамогранит, мозаика'
        except:
            category = ''

        ws.cell(row=count_def, column=9).value = category
        try:
            proiz = post['proiz']
        except:
            proiz = ''
        ws.cell(row=count_def, column=10).value = proiz
        ws.cell(row=count_def, column=11).value = ''
        try:
            link = post['link']
        except:
            link = ''
        ws.cell(row=count_def, column=12).value = link
        ws.cell(row=count_def, column=13).value = ''
        ws.cell(row=count_def, column=14).value = 'м2'
        ws.cell(row=count_def, column=15).value = ''
        ws.cell(row=count_def, column=16).value = ''
        ws.cell(row=count_def, column=17).value = 1
        ws.cell(row=count_def, column=18).value = 1
        ws.cell(row=count_def, column=19).value = ''
        ws.cell(row=count_def, column=20).value = 9999
        try:
            opis = post['opisanie']
        except:
            opis = ''
        ws.cell(row=count_def, column=21).value = opis
        ws.cell(row=count_def, column=22).value = ''
        ws.cell(row=count_def, column=23).value = ''
        ws.cell(row=count_def, column=24).value = ''
        ws.cell(row=count_def, column=25).value = ''
        try:
            contry = post['xarakt']['Страна']
        except:
            contry = ''
        ws.cell(row=count_def, column=26).value = contry
        try:
            proizvoditel = post['xarakt']['Производитель']
        except:
            proizvoditel = ''
        ws.cell(row=count_def, column=27).value = proizvoditel

        ws.cell(row=count_def, column=28).value = name
        try:
            color = ';'.join(x for x in post['color'])
        except:
            color = ''
        ws.cell(row=count_def, column=29).value = color
        try:
            size = post['xarakt']['Размер']
        except:
            size = ''

        ws.cell(row=count_def, column=30).value = size
        try:
            naznach = post['xarakt']['Помещение']
        except:
            naznach = ''
        ws.cell(row=count_def, column=31).value = naznach
        ws.cell(row=count_def, column=32).value = 'Дизайн'

        count = 0
        start_count = 26

        return True

    def create_title(self, ws):

        global_count = 0

        for count, col in enumerate(self.colums):
            ws.cell(row=2, column=global_count + 1).value = col
            ws.cell(row=2, column=global_count + 1).font = Font(bold=True)

            global_count += 1

        for count, col in enumerate(self.colums_harakt):
            ws.cell(row=2, column=global_count + 1).value = col
            ws.cell(row=2, column=global_count + 1).font = Font(bold=True)

            self.colums_checker[col] = global_count + 1

            global_count += 1

        return self.colums_checker

    def write_data(self, ws, count_def, post):

        ws.cell(row=count_def, column=1).value = ''
        ws.cell(row=count_def, column=2).value = ''
        try:
            name = post['full_name']
        except:
            name = ''
        ws.cell(row=count_def, column=3).value = name

        try:
            price = int(post['price'])
        except:
            try:
                price = post['price']
            except:
                price = ''

        ws.cell(row=count_def, column=4).value = price
        ws.cell(row=count_def, column=5).value = ''
        ws.cell(row=count_def, column=6).value = ''
        ws.cell(row=count_def, column=7).value = ''

        try:
            image = post['image']
        except:
            image = ''

        ws.cell(row=count_def, column=8).value = image

        try:
            category = 'Плитка, керамогранит, мозаика'
        except:
            category = ''

        ws.cell(row=count_def, column=9).value = category
        try:
            proiz = post['prozvoditel']
        except:
            proiz = ''
        ws.cell(row=count_def, column=10).value = proiz
        ws.cell(row=count_def, column=11).value = ''
        try:
            link = post['link']
        except:
            link = ''
        ws.cell(row=count_def, column=12).value = link
        ws.cell(row=count_def, column=13).value = ''
        try:
            ed = post['edinicha']
        except:
            ed = ''
        ws.cell(row=count_def, column=14).value = ed
        ws.cell(row=count_def, column=15).value = ''
        ws.cell(row=count_def, column=16).value = ''
        ws.cell(row=count_def, column=17).value = 1
        ws.cell(row=count_def, column=18).value = 1
        ws.cell(row=count_def, column=19).value = ''
        ws.cell(row=count_def, column=20).value = 9999
        try:
            opis = post['text']
        except:
            opis = ''
        ws.cell(row=count_def, column=21).value = opis
        ws.cell(row=count_def, column=22).value = ''
        ws.cell(row=count_def, column=23).value = ''
        ws.cell(row=count_def, column=24).value = ''
        try:
            artikl = post['artikle']
        except:
            artikl = ''
        ws.cell(row=count_def, column=25).value = artikl
        try:
            country = post['xarakt']['Страна']
        except:
            country = ''
        ws.cell(row=count_def, column=26).value = country
        ws.cell(row=count_def, column=27).value = proiz
        try:
            coll_name = post['collection']
        except:
            coll_name = ''
        ws.cell(row=count_def, column=28).value = coll_name

        try:
            color = post['xarakt']['Цвет']
        except:
            color = ''
        ws.cell(row=count_def, column=29).value = color
        try:
            size = post['xarakt']['Размер']
        except:
            size = ''

        ws.cell(row=count_def, column=30).value = size

        try:
            naznach = post['xarakt']['Помещение']
        except:
            naznach = ''
        ws.cell(row=count_def, column=31).value = naznach
        try:
            sostav_collection = post['xarakt']['Назначение']
        except:
            sostav_collection = ''
        ws.cell(row=count_def, column=32).value = sostav_collection

        try:
            picture = post['xarakt']['Рисунок']
        except:
            picture = ''
        ws.cell(row=count_def, column=33).value = picture
        try:
            poverhnost = post['xarakt']['Поверхность']
        except:
            poverhnost = ''
        ws.cell(row=count_def, column=34).value = poverhnost
        try:
            rect = post['xarakt']['Ректификат']
        except:
            rect = ''
        ws.cell(row=count_def, column=35).value = rect
        try:
            form = post['xarakt']['Форма']
        except:
            form = ''
        ws.cell(row=count_def, column=36).value = form
        try:
            width = post['xarakt']['Ширина, см']
        except:
            width = ''
        ws.cell(row=count_def, column=37).value = width
        try:
            length = post['xarakt']['Длина, см']
        except:
            length = ''
        ws.cell(row=count_def, column=38).value = length
        try:
            thickness = post['xarakt']['Толщина, мм']
        except:
            thickness = ''
        ws.cell(row=count_def, column=39).value = thickness
        try:
            count_m2 = post['xarakt']['Количество в коробке, м2']
        except:
            count_m2 = ''
        ws.cell(row=count_def, column=40).value = count_m2
        try:
            count_sh = post['xarakt']['Количество в коробке, шт.']
        except:
            count_sh = ''
        ws.cell(row=count_def, column=41).value = count_sh
        try:
            weight = post['xarakt']['Вес коробки, кг']
        except:
            weight = ''
        ws.cell(row=count_def, column=42).value = weight

        count = 0
        start_count = 33

        for key, value in post['xarakt'].items():
            # for key, value in self.colums_checker.items():
            try:
                ws.cell(row=count_def, column=self.colums_checker[key]).value = value
            except:
                continue
            # ws.cell(row=count_def + count, column=start_count).value = comment['author_comment']

            count += 1
            start_count += 1

        return True

    def itter_rows(self, ws):
        count_def = 3

        for count_post, post in enumerate(self.collect_dict):
            if post['link'] == '':
                continue

            try:
                write_coll = self.write_collections(ws, count_def, post)
            except Exception as es:
                print(f'SaveResult: ошибка write_coll {es}')

            count_def += 1

        for count_post, post in enumerate(self.good_dict):
            if post['link'] == '':
                continue

            try:
                write_data = self.write_data(ws, count_def, post)
            except Exception as es:
                print(f'SaveResult: Исключение {es}')

            count_def += 1

        return True

    def create_number_uncolums(self, ws):

        global_count = 0
        start_count = 24

        id_category = [186, 184, 9, 15, 10, 1, 45, 14, 17, 22, 96, 205, 56, 582, 571, 209, 27, 26, 590]

        for col in range(len(id_category)):
        # for col in range(len(self.colums_harakt) + 9):
            ws.cell(row=1, column=start_count + global_count).value = id_category[col]
            # ws.cell(row=1, column=start_count + global_count).value = random.randint(1111, 9999)

            global_count += 1

        return self.colums_checker

    def one_sheet(self, ws):

        self.create_number_uncolums(ws)

        name_colums_dict = self.create_title(ws)

        response_itter = self.itter_rows(ws)

        return True

    def save_file(self, filename):

        wb = Workbook()

        ws = wb.active

        result = self.one_sheet(ws)

        save_file_name = os.getcwd() + r'/files/result/tovar/tovar_' + filename + '.xlsx'

        wb.save(save_file_name)

        # filename = f'{filename}'
        #
        # wb.save(f'product_{filename}.xlsx')

        # self.save_to_json(filename, self.good_dict)

        # print(f'Сохранил \n{filename}.xlsx\n{filename}.json')
        print(f'Сохранил \n{filename}.xlsx')

        return filename

    @staticmethod
    def save_brands(links_brands):
        with open('brands.txt', 'w', encoding='utf-8') as file:
            file.write(links_brands)
            return True

    @staticmethod
    def check_brands():
        import os

        if os.path.exists('brands.txt'):
            return True
        else:
            return False

    @staticmethod
    def load_file():
        try:
            with open('brands.txt', 'r', encoding='utf-8') as file:
                files_out = file.read()
                return files_out

        except:
            return False

    @staticmethod
    def save_collection(file_name, data_dict):
        import json

        file_name = r'files/collections/' + file_name + '.json'

        try:
            with open(file_name, 'w', encoding='utf-8') as file:
                json.dump(data_dict, file, indent=4, ensure_ascii=False)

                return file_name
        except:
            return False

    @staticmethod
    def save_col_and_tovar(file_name, data_dict):
        import json

        file_name = r'files/col_product/' + file_name + '.json'

        try:
            with open(file_name, 'w', encoding='utf-8') as file:
                json.dump(data_dict, file, indent=4, ensure_ascii=False)

                return file_name
        except:
            return False

    @staticmethod
    def load_collection(file_name):
        import json

        try:
            with open(file_name, 'r', encoding='utf-8') as file:
                good_dict = json.load(file)

                return good_dict
        except:
            return False
