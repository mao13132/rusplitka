import json
import random

from openpyxl import Workbook
from openpyxl.styles import Font

from datetime import datetime


class SaveResultTovar:
    def __init__(self, good_dict):

        self.colums_checker = {}

        self.good_dict = good_dict['result']

        self.colums_harakt = good_dict['name_colums']

        self.colums = ['ID продукта', 'Код', 'Имя продукта', 'Цена', 'Старая цена', 'Комплект главные',
                       'Комплект дополнительные', 'Фото', 'Категория', 'Производитель', 'Тип товара',
                       'Ссылка на сторонний сайт', 'Алгоритм', 'Ед.Измерения', 'ID', 'PARENT_ID', 'Видимость',
                       'Видимость варианта', 'Статус товара', 'Количество', 'Описание', 'Видео', 'Документы',
                       'Гарантия', 'Артикул', 'Страна Бренда', 'Производитель', 'Коллекция', 'Цвет', 'Размер', 'Тип',
                       'Состав коллекции']

        # print()

    @staticmethod
    def save_to_json(filename, good_data):
        filename = f'{filename}.json'

        try:
            with open(filename, 'w', encoding='utf-8') as file:
                json.dump(good_data, file, indent=4, ensure_ascii=False)
        except:
            return False

        return filename

    def create_title(self, ws):

        global_count = 0

        for count, col in enumerate(self.colums):
            ws.cell(row=2, column=global_count + 1).value = col
            ws.cell(row=2, column=global_count + 1).font = Font(bold=True)

            global_count += 1

        for count, col in enumerate(self.colums_harakt):
            ws.cell(row=2, column=global_count + 1).value = col
            ws.cell(row=2, column=global_count + 1).font = Font(bold=True)

            self.colums_checker[col] = global_count + 1

            global_count += 1

        return self.colums_checker

    # def insert_contry(self, dict_in):
    #
    #     for key, value in dict_in.items():
    #         print(key)

    def write_data(self, ws, count_def, post):

        ws.cell(row=count_def, column=1).value = ''
        ws.cell(row=count_def, column=2).value = ''
        try:
            name = post['name']
        except:
            name = ''
        ws.cell(row=count_def, column=3).value = name

        try:
            price = int(post['price'])
        except:
            try:
                price = post['price']
            except:
                price = ''

        ws.cell(row=count_def, column=4).value = price
        ws.cell(row=count_def, column=5).value = ''
        ws.cell(row=count_def, column=6).value = ''
        ws.cell(row=count_def, column=7).value = ''

        try:
            image = post['image']
        except:
            image = ''

        ws.cell(row=count_def, column=8).value = image

        try:
            category = 'Плитка, керамогранит, мозаика'
        except:
            category = ''

        ws.cell(row=count_def, column=9).value = category
        try:
            proiz = post['prozvoditel']
        except:
            proiz = ''
        ws.cell(row=count_def, column=10).value = proiz
        ws.cell(row=count_def, column=11).value = ''
        try:
            link = post['link']
        except:
            link = ''
        ws.cell(row=count_def, column=12).value = link
        ws.cell(row=count_def, column=13).value = ''
        try:
            ed = post['edinicha']
        except:
            ed = ''
        ws.cell(row=count_def, column=14).value = ed
        ws.cell(row=count_def, column=15).value = ''
        ws.cell(row=count_def, column=16).value = ''
        ws.cell(row=count_def, column=17).value = 1
        ws.cell(row=count_def, column=18).value = 1
        ws.cell(row=count_def, column=19).value = ''
        ws.cell(row=count_def, column=20).value = 9999
        try:
            opis = post['text']
        except:
            opis = ''
        ws.cell(row=count_def, column=21).value = opis
        ws.cell(row=count_def, column=22).value = ''
        ws.cell(row=count_def, column=23).value = ''
        ws.cell(row=count_def, column=24).value = ''
        try:
            artikl = post['artikle']
        except:
            artikl = ''
        ws.cell(row=count_def, column=25).value = artikl
        try:
            country = post['xarakt']['Страна']
        except:
            country = ''
        ws.cell(row=count_def, column=26).value = country
        ws.cell(row=count_def, column=27).value = ''
        try:
            coll_name = post['collection']
        except:
            coll_name = ''
        ws.cell(row=count_def, column=28).value = coll_name
        ws.cell(row=count_def, column=29).value = ''
        try:
            size = post['xarakt']['Размер']
        except:
            size = ''

        ws.cell(row=count_def, column=30).value = size
        try:
            type_ = post['xarakt']['Тип']
        except:
            type_ = ''
        ws.cell(row=count_def, column=31).value = type_

        ws.cell(row=count_def, column=32).value = type_

        count = 0
        start_count = 33

        for key, value in post['xarakt'].items():
            # for key, value in self.colums_checker.items():
            try:
                ws.cell(row=count_def, column=self.colums_checker[key]).value = value
            except:
                continue
            # ws.cell(row=count_def + count, column=start_count).value = comment['author_comment']

            count += 1
            start_count += 1

        return True

    def itter_rows(self, ws):
        count_def = 3
        for count_post, post in enumerate(self.good_dict):
            if post['link'] == '':
                continue
            try:
                write_data = self.write_data(ws, count_def, post)
            except Exception as es:
                print(f'SaveResult: Исключение {es}')

            count_def += 1

        return True

    def create_number_uncolums(self, ws):

        global_count = 0
        start_count = 33

        for col in range(len(self.colums_harakt) + 5):
            ws.cell(row=1, column=start_count + global_count).value = random.randint(1111, 9999)

            global_count += 1

        return self.colums_checker

    def one_sheet(self, ws):

        self.create_number_uncolums(ws)

        name_colums_dict = self.create_title(ws)

        response_itter = self.itter_rows(ws)

        return True

    def save_file(self, filename):

        wb = Workbook()

        ws = wb.active

        result = self.one_sheet(ws)

        filename = f'{filename}'

        wb.save(f'product_{filename}.xlsx')

        # self.save_to_json(filename, self.good_dict)

        # print(f'Сохранил \n{filename}.xlsx\n{filename}.json')
        print(f'Сохранил \n{filename}.xlsx')

        return filename
